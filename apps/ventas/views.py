import json
from decimal import Decimal
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.utils import timezone
from django.db.models import Sum
from .models import Venta, ItemVenta, CorteCaja
from .forms import VentaForm
from apps.barberos.models import Barbero
from apps.servicios.models import Servicio
from apps.inventario.models import Producto
from apps.accounts.utils import no_barbero

@login_required
def lista_ventas(request):
    hoy = timezone.now().date()
    ventas = Venta.objects.select_related('cliente', 'barbero__usuario').order_by('-fecha')
    if request.user.es_barbero:
        try:
            ventas = ventas.filter(barbero=request.user.perfil_barbero)
        except Exception:
            ventas = ventas.none()
    ventas = ventas[:100]
    total_hoy_qs = Venta.objects.filter(fecha__date=hoy)
    if request.user.es_barbero:
        try:
            total_hoy_qs = total_hoy_qs.filter(barbero=request.user.perfil_barbero)
        except Exception:
            total_hoy_qs = total_hoy_qs.none()
    total_hoy = total_hoy_qs.aggregate(t=Sum('total'))['t'] or 0
    return render(request, 'ventas/lista.html', {'ventas': ventas, 'total_hoy': total_hoy})

@login_required
def nueva_venta(request):
    servicios = Servicio.objects.filter(activo=True).select_related('categoria').order_by('nombre')
    productos = Producto.objects.filter(activo=True).order_by('nombre')
    barberos  = Barbero.objects.filter(activo=True).select_related('usuario')

    # Barbero propio si el usuario tiene ese rol
    barbero_propio = None
    if request.user.es_barbero:
        try:
            barbero_propio = request.user.perfil_barbero
        except Exception:
            pass

    # Pre-llenado desde cita (GET ?cita=<pk>)
    prefill_json = 'null'
    cita_pk = request.GET.get('cita')
    if cita_pk:
        try:
            from apps.citas.models import Cita
            c = Cita.objects.select_related('cliente', 'barbero__usuario', 'servicio').get(pk=cita_pk)
            prefill = {
                'cita_id':           c.pk,
                'cliente_id':        c.cliente_id,
                'cliente_nombre':    c.cliente.nombre,
                'cliente_telefono':  c.cliente.telefono or '',
                'barbero_id':        c.barbero_id,
                'servicio_id':       c.servicio_id or '',
                'servicio_nombre':   c.servicio.nombre if c.servicio else '',
                'servicio_precio':   float(c.servicio.precio) if c.servicio else 0,
            }
            prefill_json = json.dumps(prefill)
        except Exception:
            pass

    servicios_json = json.dumps([
        {'id': s.pk, 'nombre': s.nombre, 'precio': float(s.precio), 'duracion': s.duracion_minutos}
        for s in servicios
    ])
    productos_json = json.dumps([
        {'id': p.pk, 'nombre': p.nombre, 'precio': float(p.precio_venta), 'stock': p.stock_actual}
        for p in productos
    ])

    if request.method == 'POST':
        # Para rol barbero se asigna automáticamente su propio perfil
        if request.user.es_barbero and barbero_propio:
            barbero_id = barbero_propio.pk
        else:
            barbero_id = request.POST.get('barbero') or None
        cliente_id  = request.POST.get('cliente') or None
        cita_id     = request.POST.get('cita') or None
        metodo_pago = request.POST.get('metodo_pago', 'efectivo')
        descuento   = request.POST.get('descuento', '0') or '0'
        descuento   = Decimal(descuento).quantize(Decimal('0.01'))
        notas       = request.POST.get('notas', '')

        ctx_err = {'servicios': servicios, 'productos': productos, 'barberos': barberos,
                   'servicios_json': servicios_json, 'productos_json': productos_json,
                   'hoy': timezone.now().date().isoformat()}

        if not barbero_id:
            messages.error(request, 'Selecciona un barbero.')
            return render(request, 'ventas/nueva.html', ctx_err)

        tipos      = request.POST.getlist('item_tipo')
        ids        = request.POST.getlist('item_id')
        cantidades = request.POST.getlist('item_cantidad')
        precios    = request.POST.getlist('item_precio')
        descs      = request.POST.getlist('item_desc')

        if not any(ids):
            messages.error(request, 'Agrega al menos un servicio o producto.')
            return render(request, 'ventas/nueva.html', ctx_err)

        if cita_id and Venta.objects.filter(cita_id=cita_id).exists():
            messages.error(request, 'Esta cita ya tiene una venta registrada.')
            return render(request, 'ventas/nueva.html', ctx_err)

        try:
            venta = Venta(
                barbero_id   = barbero_id,
                cliente_id   = cliente_id or None,
                metodo_pago  = metodo_pago,
                descuento    = descuento,
                notas        = notas,
                atendido_por = request.user,
            )
            if cita_id:
                venta.cita_id = cita_id
            venta.save()

            for i in range(len(tipos)):
                if not tipos[i] or not ids[i]:
                    continue
                desc = descs[i] if i < len(descs) else ''
                cant = int(cantidades[i]) if i < len(cantidades) and cantidades[i] else 1
                precio = Decimal(precios[i]).quantize(Decimal('0.01')) if i < len(precios) and precios[i] else Decimal('0')

                item = ItemVenta(
                    venta       = venta,
                    tipo        = tipos[i],
                    descripcion = desc,
                    cantidad    = cant,
                    precio_unitario = precio,
                )
                if tipos[i] == 'servicio':
                    item.servicio = Servicio.objects.filter(pk=ids[i]).first()
                else:
                    item.producto = Producto.objects.filter(pk=ids[i]).first()
                item.save()

            venta.calcular_total()
            # Marcar cita como atendida si se vinculó
            if cita_id:
                from apps.citas.models import Cita
                Cita.objects.filter(pk=cita_id).update(estado='atendido', precio_cobrado=venta.total)

            messages.success(request, f'✔ Venta #{venta.pk} registrada — Total: Bs. {venta.total}')
            return redirect('ventas:detalle', pk=venta.pk)

        except Exception as e:
            messages.error(request, f'Error al registrar venta: {e}')

    return render(request, 'ventas/nueva.html', {
        'servicios':      servicios,
        'productos':      productos,
        'barberos':       barberos,
        'barbero_propio': barbero_propio,
        'hoy':            timezone.now().date().isoformat(),
        'servicios_json': servicios_json,
        'productos_json': productos_json,
        'prefill_json':   prefill_json,
    })

@login_required
def detalle_venta(request, pk):
    venta = get_object_or_404(Venta, pk=pk)
    if request.user.es_barbero:
        try:
            if venta.barbero != request.user.perfil_barbero:
                messages.error(request, 'Solo puedes ver tus propias ventas.')
                return redirect('ventas:lista')
        except Exception:
            messages.error(request, 'Sin acceso.')
            return redirect('citas:dashboard')
    return render(request, 'ventas/detalle.html', {'venta': venta})

def _resumen_ventas(qs):
    ef = qs.filter(metodo_pago='efectivo').aggregate(t=Sum('total'))['t'] or 0
    qr = qs.filter(metodo_pago='qr').aggregate(t=Sum('total'))['t'] or 0
    ta = qs.filter(metodo_pago='tarjeta').aggregate(t=Sum('total'))['t'] or 0
    return {'efectivo': ef, 'qr': qr, 'tarjeta': ta, 'total': ef + qr + ta,
            'cantidad': qs.count()}

@login_required
@no_barbero
def corte_caja(request):
    hoy = timezone.now().date()
    periodo = request.GET.get('periodo', 'dia')

    if periodo == 'mes':
        ventas = Venta.objects.filter(fecha__year=hoy.year, fecha__month=hoy.month)
        titulo = f"{hoy.strftime('%B %Y').upper()}"
    elif periodo == 'anio':
        ventas = Venta.objects.filter(fecha__year=hoy.year)
        titulo = str(hoy.year)
    else:
        ventas = Venta.objects.filter(fecha__date=hoy)
        titulo = hoy.strftime('%d/%m/%Y')
        periodo = 'dia'

    if request.user.es_barbero:
        try:
            ventas = ventas.filter(barbero=request.user.perfil_barbero)
        except Exception:
            ventas = ventas.none()

    ventas = ventas.select_related('cliente', 'barbero__usuario').order_by('-fecha')
    resumen = _resumen_ventas(ventas)

    # Resumen por barbero
    from django.db.models import Count
    por_barbero = (ventas.values('barbero__usuario__first_name', 'barbero__usuario__last_name',
                                 'barbero__usuario__username')
                         .annotate(total=Sum('total'), cantidad=Count('id'))
                         .order_by('-total'))

    corte_existente = CorteCaja.objects.filter(fecha=hoy).first()
    return render(request, 'ventas/corte_caja.html', {
        'resumen': resumen, 'hoy': hoy, 'titulo': titulo,
        'ventas': ventas, 'corte': corte_existente,
        'periodo': periodo, 'por_barbero': por_barbero,
    })

@login_required
@no_barbero
def realizar_corte(request):
    if request.method == 'POST':
        hoy = timezone.now().date()
        ventas_hoy = Venta.objects.filter(fecha__date=hoy)
        ef = ventas_hoy.filter(metodo_pago='efectivo').aggregate(t=Sum('total'))['t'] or 0
        qr = ventas_hoy.filter(metodo_pago='qr').aggregate(t=Sum('total'))['t'] or 0
        ta = ventas_hoy.filter(metodo_pago='tarjeta').aggregate(t=Sum('total'))['t'] or 0
        corte, creado = CorteCaja.objects.get_or_create(fecha=hoy, defaults={
            'efectivo': ef, 'qr': qr, 'tarjeta': ta, 'total': ef+qr+ta,
            'cerrado_por': request.user, 'notas': request.POST.get('notas', '')
        })
        if creado:
            messages.success(request, f'Corte de caja realizado. Total: Bs. {corte.total}')
        else:
            messages.warning(request, 'Ya existe un corte para hoy.')
    return redirect('ventas:corte_caja')


@login_required
@no_barbero
def exportar_excel(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from django.http import HttpResponse

    hoy = timezone.now().date()
    periodo = request.GET.get('periodo', 'dia')

    if periodo == 'mes':
        ventas = Venta.objects.filter(fecha__year=hoy.year, fecha__month=hoy.month)
        nombre_periodo = hoy.strftime('%B_%Y')
    elif periodo == 'anio':
        ventas = Venta.objects.filter(fecha__year=hoy.year)
        nombre_periodo = str(hoy.year)
    else:
        ventas = Venta.objects.filter(fecha__date=hoy)
        nombre_periodo = hoy.strftime('%d-%m-%Y')

    if request.user.es_barbero:
        try:
            ventas = ventas.filter(barbero=request.user.perfil_barbero)
        except Exception:
            ventas = ventas.none()

    ventas = ventas.select_related('cliente', 'barbero__usuario').prefetch_related('items').order_by('-fecha')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Ventas'

    # Estilos
    COLOR_HEADER = '1a1a2e'
    COLOR_ACCENT = 'c8a951'
    COLOR_ALT    = '16213e'

    hdr_font  = Font(name='Calibri', bold=True, color=COLOR_ACCENT, size=11)
    hdr_fill  = PatternFill('solid', fgColor=COLOR_HEADER)
    alt_fill  = PatternFill('solid', fgColor=COLOR_ALT)
    thin      = Border(
        left=Side(style='thin', color='2a2a4a'),
        right=Side(style='thin', color='2a2a4a'),
        top=Side(style='thin', color='2a2a4a'),
        bottom=Side(style='thin', color='2a2a4a'),
    )
    center    = Alignment(horizontal='center', vertical='center')
    right_al  = Alignment(horizontal='right', vertical='center')

    # Título
    ws.merge_cells('A1:I1')
    ws['A1'] = f'REPORTE DE VENTAS — {nombre_periodo.upper()}'
    ws['A1'].font = Font(name='Calibri', bold=True, color=COLOR_ACCENT, size=14)
    ws['A1'].fill = PatternFill('solid', fgColor='0d0d1a')
    ws['A1'].alignment = center
    ws.row_dimensions[1].height = 30

    # Encabezados
    headers = ['#', 'Fecha', 'Hora', 'Cliente', 'Barbero', 'Método', 'Ítems', 'Descuento', 'Total']
    ws.append([])  # fila 2 vacía
    ws.append(headers)  # fila 3
    for col, _ in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col)
        cell.font  = hdr_font
        cell.fill  = hdr_fill
        cell.border = thin
        cell.alignment = center
    ws.row_dimensions[3].height = 20

    # Datos
    for i, v in enumerate(ventas):
        row_idx = 4 + i
        barbero_nombre = ''
        if v.barbero and v.barbero.usuario:
            u = v.barbero.usuario
            barbero_nombre = u.get_full_name() or u.username
        items_desc = ', '.join(it.descripcion for it in v.items.all())

        row_data = [
            v.pk,
            v.fecha.strftime('%d/%m/%Y'),
            v.fecha.strftime('%H:%M'),
            v.cliente.nombre if v.cliente else '—',
            barbero_nombre,
            v.get_metodo_pago_display(),
            items_desc,
            float(v.descuento),
            float(v.total),
        ]
        ws.append(row_data)
        fill = alt_fill if i % 2 else None
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.border = thin
            if fill:
                cell.fill = fill
            if col in (8, 9):
                cell.alignment = right_al
                cell.number_format = '#,##0.00'
            else:
                cell.alignment = Alignment(vertical='center')

    # Fila de totales
    total_row = 4 + len(ventas)
    ws.cell(row=total_row, column=8, value='TOTAL')
    ws.cell(row=total_row, column=8).font = Font(bold=True, color=COLOR_ACCENT)
    ws.cell(row=total_row, column=8).alignment = right_al
    total_val = ventas.aggregate(t=Sum('total'))['t'] or 0
    ws.cell(row=total_row, column=9, value=float(total_val))
    ws.cell(row=total_row, column=9).font = Font(bold=True, color=COLOR_ACCENT)
    ws.cell(row=total_row, column=9).number_format = '#,##0.00'
    ws.cell(row=total_row, column=9).alignment = right_al

    # Ancho de columnas
    anchos = [6, 12, 8, 22, 20, 12, 40, 12, 12]
    for col, ancho in enumerate(anchos, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = ancho

    # Hoja resumen
    ws2 = wb.create_sheet('Resumen')
    ws2.column_dimensions['A'].width = 22
    ws2.column_dimensions['B'].width = 16
    resumen_data = [
        ('Período', nombre_periodo),
        ('Total ventas', int(ventas.count())),
        ('Efectivo (Bs.)', float(ventas.filter(metodo_pago='efectivo').aggregate(t=Sum('total'))['t'] or 0)),
        ('QR / Transfer (Bs.)', float(ventas.filter(metodo_pago='qr').aggregate(t=Sum('total'))['t'] or 0)),
        ('Tarjeta (Bs.)', float(ventas.filter(metodo_pago='tarjeta').aggregate(t=Sum('total'))['t'] or 0)),
        ('TOTAL (Bs.)', float(total_val)),
    ]
    for r, (label, val) in enumerate(resumen_data, 1):
        ws2.cell(row=r, column=1, value=label).font = Font(bold=True, color='c8a951')
        ws2.cell(row=r, column=2, value=val)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="ventas_{nombre_periodo}.xlsx"'
    wb.save(response)
    return response
