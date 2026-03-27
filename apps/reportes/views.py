from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, Count, Avg
from django.utils import timezone
from datetime import timedelta
from apps.ventas.models import Venta, ItemVenta
from apps.citas.models import Cita
from apps.clientes.models import Cliente
from apps.barberos.models import Barbero


def solo_admin(view_func):
    from functools import wraps
    from django.http import HttpResponseForbidden
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated or request.user.rol not in ['admin']:
            return HttpResponseForbidden('Solo administradores.')
        return view_func(request, *args, **kwargs)
    return wrapper


@login_required
@solo_admin
def reporte_general(request):
    hoy = timezone.now().date()
    inicio_mes = hoy.replace(day=1)
    ventas_mes = Venta.objects.filter(fecha__date__gte=inicio_mes)
    ingresos_mes = ventas_mes.aggregate(t=Sum('total'))['t'] or 0
    total_citas_mes = Cita.objects.filter(fecha_hora__date__gte=inicio_mes).count()
    no_shows = Cita.objects.filter(fecha_hora__date__gte=inicio_mes, estado='no_show').count()
    clientes_nuevos = Cliente.objects.filter(fecha_registro__date__gte=inicio_mes).count()
    context = {
        'ingresos_mes': ingresos_mes, 'total_citas_mes': total_citas_mes,
        'no_shows': no_shows, 'clientes_nuevos': clientes_nuevos,
    }
    return render(request, 'reportes/general.html', context)


@login_required
@solo_admin
def reporte_ingresos(request):
    dias = int(request.GET.get('dias', 30))
    desde = timezone.now().date() - timedelta(days=dias)
    ventas = Venta.objects.filter(fecha__date__gte=desde).order_by('fecha')
    por_dia = {}
    for v in ventas:
        dia = v.fecha.date().isoformat()
        por_dia[dia] = por_dia.get(dia, 0) + float(v.total)
    return render(request, 'reportes/ingresos.html', {'por_dia': por_dia, 'dias': dias, 'ventas': ventas})


@login_required
@solo_admin
def reporte_servicios(request):
    dias = int(request.GET.get('dias', 30))
    desde = timezone.now().date() - timedelta(days=dias)
    servicios = ItemVenta.objects.filter(
        tipo='servicio', venta__fecha__date__gte=desde
    ).values('servicio__nombre').annotate(
        total_vendido=Count('id'), ingresos=Sum('precio_unitario')
    ).order_by('-total_vendido')[:10]
    return render(request, 'reportes/servicios.html', {'servicios': servicios, 'dias': dias})


@login_required
@solo_admin
def reporte_barberos(request):
    from decimal import Decimal
    hoy = timezone.now().date()
    periodo = request.GET.get('periodo', 'mes')

    if periodo == 'dia':
        ventas_qs = Venta.objects.filter(fecha__date=hoy)
        titulo = hoy.strftime('%d/%m/%Y')
    elif periodo == 'anio':
        ventas_qs = Venta.objects.filter(fecha__year=hoy.year)
        titulo = str(hoy.year)
    else:
        ventas_qs = Venta.objects.filter(fecha__year=hoy.year, fecha__month=hoy.month)
        titulo = hoy.strftime('%B %Y').upper()
        periodo = 'mes'

    # Agregar por barbero incluyendo su % de comisión
    por_barbero = (ventas_qs
        .values(
            'barbero__pk',
            'barbero__usuario__first_name',
            'barbero__usuario__last_name',
            'barbero__usuario__username',
            'barbero__comision_porcentaje',
        )
        .annotate(total_ventas=Count('id'), ingresos=Sum('total'))
        .order_by('-ingresos'))

    # Calcular comisiones en Python y armar lista
    barberos_data = []
    total_ingresos = Decimal('0')
    total_comisiones = Decimal('0')
    max_ingresos = Decimal('1')  # evitar división por cero para la barra visual

    for b in por_barbero:
        ingresos = b['ingresos'] or Decimal('0')
        comision_pct = b['barbero__comision_porcentaje'] or Decimal('0')
        comision = (ingresos * comision_pct / 100).quantize(Decimal('0.01'))
        neto = ingresos - comision
        nombre = (
            f"{b['barbero__usuario__first_name']} {b['barbero__usuario__last_name']}".strip()
            or b['barbero__usuario__username']
        )
        barberos_data.append({
            'pk': b['barbero__pk'],
            'nombre': nombre,
            'total_ventas': b['total_ventas'],
            'ingresos': ingresos,
            'comision_pct': comision_pct,
            'comision': comision,
            'neto': neto,
        })
        total_ingresos += ingresos
        total_comisiones += comision
        if ingresos > max_ingresos:
            max_ingresos = ingresos

    # Añadir porcentaje de barra visual
    for b in barberos_data:
        b['barra_pct'] = int(b['ingresos'] / max_ingresos * 100)

    return render(request, 'reportes/barberos.html', {
        'barberos_data': barberos_data,
        'periodo': periodo,
        'titulo': titulo,
        'hoy': hoy,
        'total_ingresos': total_ingresos,
        'total_comisiones': total_comisiones,
        'total_neto': total_ingresos - total_comisiones,
    })


@login_required
@solo_admin
def exportar_barberos_excel(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from django.http import HttpResponse
    from decimal import Decimal

    hoy = timezone.now().date()
    periodo = request.GET.get('periodo', 'mes')

    if periodo == 'dia':
        ventas_qs = Venta.objects.filter(fecha__date=hoy)
        nombre_periodo = hoy.strftime('%d-%m-%Y')
        titulo_periodo = hoy.strftime('%d/%m/%Y')
    elif periodo == 'anio':
        ventas_qs = Venta.objects.filter(fecha__year=hoy.year)
        nombre_periodo = str(hoy.year)
        titulo_periodo = str(hoy.year)
    else:
        ventas_qs = Venta.objects.filter(fecha__year=hoy.year, fecha__month=hoy.month)
        nombre_periodo = hoy.strftime('%B_%Y')
        titulo_periodo = hoy.strftime('%B %Y').upper()

    por_barbero = (ventas_qs
        .values(
            'barbero__pk',
            'barbero__usuario__first_name',
            'barbero__usuario__last_name',
            'barbero__usuario__username',
            'barbero__comision_porcentaje',
        )
        .annotate(total_ventas=Count('id'), ingresos=Sum('total'))
        .order_by('-ingresos'))

    barberos_data = []
    total_ingresos = Decimal('0')
    total_comisiones = Decimal('0')
    for b in por_barbero:
        ingresos = b['ingresos'] or Decimal('0')
        comision_pct = b['barbero__comision_porcentaje'] or Decimal('0')
        comision = (ingresos * comision_pct / 100).quantize(Decimal('0.01'))
        neto = ingresos - comision
        nombre = (
            f"{b['barbero__usuario__first_name']} {b['barbero__usuario__last_name']}".strip()
            or b['barbero__usuario__username']
        )
        barberos_data.append({
            'nombre': nombre,
            'total_ventas': b['total_ventas'],
            'ingresos': ingresos,
            'comision_pct': comision_pct,
            'comision': comision,
            'neto': neto,
        })
        total_ingresos += ingresos
        total_comisiones += comision

    # ── Estilos ──
    COLOR_BG     = '0d0d1a'
    COLOR_HDR    = '1a1a2e'
    COLOR_ALT    = '16213e'
    COLOR_GOLD   = 'c8a951'
    COLOR_GREEN  = '2d6a2d'
    COLOR_RED    = '8b1a1a'

    hdr_font  = Font(name='Calibri', bold=True, color=COLOR_GOLD, size=11)
    hdr_fill  = PatternFill('solid', fgColor=COLOR_HDR)
    alt_fill  = PatternFill('solid', fgColor=COLOR_ALT)
    thin      = Border(
        left=Side(style='thin', color='2a2a4a'), right=Side(style='thin', color='2a2a4a'),
        top=Side(style='thin', color='2a2a4a'), bottom=Side(style='thin', color='2a2a4a'),
    )
    center = Alignment(horizontal='center', vertical='center')
    right  = Alignment(horizontal='right',  vertical='center')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Rendimiento Barberos'

    # Título
    ws.merge_cells('A1:G1')
    ws['A1'] = f'REPORTE DE BARBEROS — {titulo_periodo}'
    ws['A1'].font  = Font(name='Calibri', bold=True, color=COLOR_GOLD, size=14)
    ws['A1'].fill  = PatternFill('solid', fgColor=COLOR_BG)
    ws['A1'].alignment = center
    ws.row_dimensions[1].height = 30

    headers = ['#', 'Barbero', 'Ventas', 'Ingresos Brutos (Bs.)', '% Comisión', 'Comisión Barbero (Bs.)', 'Neto Local (Bs.)']
    ws.append([])   # fila 2 vacía
    ws.append(headers)  # fila 3
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=3, column=col)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.border = thin
        cell.alignment = center
    ws.row_dimensions[3].height = 20

    for i, b in enumerate(barberos_data):
        row_idx = 4 + i
        row_data = [
            i + 1,
            b['nombre'],
            b['total_ventas'],
            float(b['ingresos']),
            float(b['comision_pct']),
            float(b['comision']),
            float(b['neto']),
        ]
        ws.append(row_data)
        fill = alt_fill if i % 2 else None
        for col, _ in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.border = thin
            if fill:
                cell.fill = fill
            if col in (4, 6, 7):
                cell.number_format = '#,##0.00'
                cell.alignment = right
            elif col == 5:
                cell.number_format = '0.00"%"'
                cell.alignment = center
            else:
                cell.alignment = center if col == 1 else Alignment(vertical='center')

    # Fila de totales
    total_row = 4 + len(barberos_data)
    ws.cell(row=total_row, column=2, value='TOTALES').font = Font(bold=True, color=COLOR_GOLD)
    ws.cell(row=total_row, column=4, value=float(total_ingresos)).number_format = '#,##0.00'
    ws.cell(row=total_row, column=4).font  = Font(bold=True, color=COLOR_GOLD)
    ws.cell(row=total_row, column=4).alignment = right
    ws.cell(row=total_row, column=6, value=float(total_comisiones)).number_format = '#,##0.00'
    ws.cell(row=total_row, column=6).font  = Font(bold=True, color=COLOR_GOLD)
    ws.cell(row=total_row, column=6).alignment = right
    ws.cell(row=total_row, column=7, value=float(total_ingresos - total_comisiones)).number_format = '#,##0.00'
    ws.cell(row=total_row, column=7).font  = Font(bold=True, color=COLOR_GOLD)
    ws.cell(row=total_row, column=7).alignment = right

    # Anchos
    for col, ancho in zip(range(1, 8), [5, 28, 10, 22, 14, 24, 20]):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = ancho

    # Hoja Resumen
    ws2 = wb.create_sheet('Resumen')
    ws2.column_dimensions['A'].width = 26
    ws2.column_dimensions['B'].width = 18
    for r, (label, val) in enumerate([
        ('Período', titulo_periodo),
        ('Total Ingresos (Bs.)', float(total_ingresos)),
        ('Total Comisiones (Bs.)', float(total_comisiones)),
        ('Neto Local (Bs.)', float(total_ingresos - total_comisiones)),
        ('Total Barberos', len(barberos_data)),
    ], 1):
        ws2.cell(row=r, column=1, value=label).font = Font(bold=True, color=COLOR_GOLD)
        ws2.cell(row=r, column=2, value=val)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="barberos_{nombre_periodo}.xlsx"'
    wb.save(response)
    return response


@login_required
@solo_admin
def reporte_historial(request):
    from decimal import Decimal
    from datetime import date
    hoy = timezone.now().date()

    # ── Período ────────────────────────────────────────────────────────────────
    periodo = request.GET.get('periodo', 'mes')
    barbero_pk = request.GET.get('barbero', '')
    fecha_desde = request.GET.get('desde', '')
    fecha_hasta = request.GET.get('hasta', '')

    if periodo == 'dia':
        desde = hasta = hoy
        titulo_periodo = hoy.strftime('%d/%m/%Y')
    elif periodo == 'anio':
        desde = hoy.replace(month=1, day=1)
        hasta = hoy.replace(month=12, day=31)
        titulo_periodo = str(hoy.year)
    elif periodo == 'custom' and fecha_desde and fecha_hasta:
        from datetime import datetime
        try:
            desde = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
            hasta = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
            titulo_periodo = f'{desde.strftime("%d/%m/%Y")} — {hasta.strftime("%d/%m/%Y")}'
        except ValueError:
            desde = hoy.replace(day=1)
            hasta = hoy
            titulo_periodo = hoy.strftime('%B %Y').upper()
            periodo = 'mes'
    else:
        desde = hoy.replace(day=1)
        hasta = hoy
        titulo_periodo = hoy.strftime('%B %Y').upper()
        periodo = 'mes'

    # ── Queryset base ──────────────────────────────────────────────────────────
    ventas_qs = (Venta.objects
                 .filter(fecha__date__gte=desde, fecha__date__lte=hasta)
                 .select_related('barbero__usuario', 'cliente', 'cita')
                 .prefetch_related('items__servicio', 'items__producto')
                 .order_by('-fecha'))

    if barbero_pk:
        ventas_qs = ventas_qs.filter(barbero__pk=barbero_pk)

    # ── Enriquecer cada venta con comisión ─────────────────────────────────────
    ventas_data = []
    total_ingresos  = Decimal('0')
    total_comisiones = Decimal('0')

    for v in ventas_qs:
        comision_pct = (v.barbero.comision_porcentaje if v.barbero else Decimal('0')) or Decimal('0')
        comision     = (v.total * comision_pct / 100).quantize(Decimal('0.01'))
        neto         = v.total - comision
        nombre_barbero = (
            v.barbero.usuario.get_full_name().strip() or v.barbero.usuario.username
            if v.barbero else '—'
        )
        items = list(v.items.all())
        ventas_data.append({
            'pk':           v.pk,
            'fecha':        v.fecha,
            'barbero':      nombre_barbero,
            'barbero_pk':   v.barbero_id,
            'cliente':      v.cliente.nombre if v.cliente else '—',
            'metodo_pago':  v.get_metodo_pago_display(),
            'subtotal':     v.subtotal,
            'descuento':    v.descuento,
            'total':        v.total,
            'comision_pct': comision_pct,
            'comision':     comision,
            'neto':         neto,
            'items':        items,
            'cita_pk':      v.cita_id,
        })
        total_ingresos   += v.total
        total_comisiones += comision

    barberos = Barbero.objects.filter(activo=True).select_related('usuario')

    return render(request, 'reportes/historial.html', {
        'ventas_data':      ventas_data,
        'periodo':          periodo,
        'titulo_periodo':   titulo_periodo,
        'hoy':              hoy,
        'barberos':         barberos,
        'barbero_pk':       barbero_pk,
        'fecha_desde':      fecha_desde,
        'fecha_hasta':      fecha_hasta,
        'total_ingresos':   total_ingresos,
        'total_comisiones': total_comisiones,
        'total_neto':       total_ingresos - total_comisiones,
        'total_ventas':     len(ventas_data),
    })


@login_required
@solo_admin
def exportar_historial_excel(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from django.http import HttpResponse
    from decimal import Decimal
    from datetime import datetime

    hoy = timezone.now().date()
    periodo    = request.GET.get('periodo', 'mes')
    barbero_pk = request.GET.get('barbero', '')
    fecha_desde = request.GET.get('desde', '')
    fecha_hasta = request.GET.get('hasta', '')

    if periodo == 'dia':
        desde = hasta = hoy
        nombre_archivo = hoy.strftime('%d-%m-%Y')
    elif periodo == 'anio':
        desde = hoy.replace(month=1, day=1)
        hasta = hoy.replace(month=12, day=31)
        nombre_archivo = str(hoy.year)
    elif periodo == 'custom' and fecha_desde and fecha_hasta:
        try:
            desde = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
            hasta = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
            nombre_archivo = f'{fecha_desde}_{fecha_hasta}'
        except ValueError:
            desde = hoy.replace(day=1)
            hasta = hoy
            nombre_archivo = hoy.strftime('%B_%Y')
    else:
        desde = hoy.replace(day=1)
        hasta = hoy
        nombre_archivo = hoy.strftime('%B_%Y')

    ventas_qs = (Venta.objects
                 .filter(fecha__date__gte=desde, fecha__date__lte=hasta)
                 .select_related('barbero__usuario', 'cliente')
                 .prefetch_related('items__servicio', 'items__producto')
                 .order_by('-fecha'))
    if barbero_pk:
        ventas_qs = ventas_qs.filter(barbero__pk=barbero_pk)

    # Estilos
    COLOR_BG   = '0d0d1a'
    COLOR_HDR  = '1a1a2e'
    COLOR_ALT  = '16213e'
    COLOR_GOLD = 'c8a951'
    hdr_font = Font(name='Calibri', bold=True, color=COLOR_GOLD, size=11)
    hdr_fill = PatternFill('solid', fgColor=COLOR_HDR)
    alt_fill = PatternFill('solid', fgColor=COLOR_ALT)
    thin = Border(
        left=Side(style='thin', color='2a2a4a'), right=Side(style='thin', color='2a2a4a'),
        top=Side(style='thin', color='2a2a4a'), bottom=Side(style='thin', color='2a2a4a'),
    )
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    right  = Alignment(horizontal='right',  vertical='center')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Historial Ventas'

    # Título
    ws.merge_cells('A1:K1')
    ws['A1'] = f'HISTORIAL DE VENTAS — {nombre_archivo.replace("_", " ").upper()}'
    ws['A1'].font      = Font(name='Calibri', bold=True, color=COLOR_GOLD, size=14)
    ws['A1'].fill      = PatternFill('solid', fgColor=COLOR_BG)
    ws['A1'].alignment = center
    ws.row_dimensions[1].height = 30

    headers = ['#', 'Fecha', 'Barbero', 'Cliente', 'Método Pago',
               'Ítems', 'Subtotal (Bs.)', 'Descuento (Bs.)',
               'Total (Bs.)', '% Com.', 'Comisión (Bs.)']
    ws.append([])
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=3, column=col)
        c.font = hdr_font; c.fill = hdr_fill; c.border = thin; c.alignment = center
    ws.row_dimensions[3].height = 22

    total_ingresos = total_comisiones = Decimal('0')
    for i, v in enumerate(ventas_qs):
        comision_pct = (v.barbero.comision_porcentaje if v.barbero else Decimal('0')) or Decimal('0')
        comision     = (v.total * comision_pct / 100).quantize(Decimal('0.01'))
        nombre_barbero = (
            v.barbero.usuario.get_full_name().strip() or v.barbero.usuario.username
            if v.barbero else '—'
        )
        items_txt = '; '.join(
            f'{it.descripcion} x{it.cantidad}' for it in v.items.all()
        ) or '—'
        row_data = [
            i + 1,
            v.fecha.strftime('%d/%m/%Y %H:%M'),
            nombre_barbero,
            v.cliente.nombre if v.cliente else '—',
            v.get_metodo_pago_display(),
            items_txt,
            float(v.subtotal),
            float(v.descuento),
            float(v.total),
            float(comision_pct),
            float(comision),
        ]
        ws.append(row_data)
        row_idx = 4 + i
        fill = alt_fill if i % 2 else None
        for col, _ in enumerate(row_data, 1):
            c = ws.cell(row=row_idx, column=col)
            c.border = thin
            if fill: c.fill = fill
            if col in (7, 8, 9, 11):
                c.number_format = '#,##0.00'; c.alignment = right
            elif col == 10:
                c.number_format = '0.00"%"'; c.alignment = center
            else:
                c.alignment = center if col in (1, 5) else Alignment(vertical='center', wrap_text=True)
        total_ingresos   += v.total
        total_comisiones += comision

    # Totales
    tr = 4 + ventas_qs.count()
    ws.cell(row=tr, column=3, value='TOTALES').font = Font(bold=True, color=COLOR_GOLD)
    for col, val in [(9, float(total_ingresos)), (11, float(total_comisiones))]:
        c = ws.cell(row=tr, column=col, value=val)
        c.font = Font(bold=True, color=COLOR_GOLD)
        c.number_format = '#,##0.00'
        c.alignment = right

    for col, ancho in zip(range(1, 12), [5, 18, 22, 20, 15, 40, 16, 16, 14, 8, 16]):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = ancho

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="historial_{nombre_archivo}.xlsx"'
    wb.save(response)
    return response


@login_required
@solo_admin
def reporte_clientes(request):
    clientes = Cliente.objects.annotate(
        total_citas=Count('citas'), visitas=Count('citas', filter=__import__('django.db.models', fromlist=['Q']).Q(citas__estado='atendido'))
    ).filter(activo=True).order_by('-total_citas')[:20]
    return render(request, 'reportes/clientes.html', {'clientes': clientes})
